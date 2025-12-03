import customtkinter as ctk
import tkinter as tk
from tkinter import ttk, filedialog
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from matplotlib.figure import Figure
import threading
import time
import openpyxl
from openpyxl.styles import PatternFill, Font

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

class Perceptron3DApp:
    def __init__(self):
        self.root = ctk.CTk()
        self.root.title("3 változós AND Perceptron tanítása – Interaktív 3D UI")
        self.root.geometry("1400x900")

        # === Paraméterek ===
        self.w1 = tk.DoubleVar(value=0.7)
        self.w2 = tk.DoubleVar(value=0.2)
        self.w3 = tk.DoubleVar(value=0.8)
        self.b = tk.DoubleVar(value=0.10)
        self.lr = tk.DoubleVar(value=0.2)
        self.max_iter = tk.IntVar(value=100)
        self.running = False
        self.converged = False

        self.data = [
            [0, 0, 0, 0], [0, 0, 1, 0], [0, 1, 0, 0], [0, 1, 1, 0],
            [1, 0, 0, 0], [1, 0, 1, 0], [1, 1, 0, 0], [1, 1, 1, 1]
        ]

        self.iteration_records = []  # minden iteráció adatainak tárolása

        self.setup_ui()
        self.update_plot()

    def step_function(self, net):
        return 1 if net >= 0 else 0

    def setup_ui(self):
        # === Bal oldali vezérlőpanel ===
        control_frame = ctk.CTkFrame(self.root, width=350, corner_radius=15)
        control_frame.pack(side="left", fill="y", padx=10, pady=10)
        control_frame.pack_propagate(False)

        ctk.CTkLabel(control_frame, text="Perceptron Tanítás – 3D AND", font=("Roboto", 20, "bold")).pack(pady=20)

        # Paraméterek
        params = [
            ("w1 (súly 1)", self.w1),
            ("w2 (súly 2)", self.w2),
            ("w3 (súly 3)", self.w3),
            ("bias (b)", self.b),
            ("Tanulási ráta η", self.lr),
            ("Max iterációk", self.max_iter)
        ]

        self.entries = {}
        for text, var in params:
            frame = ctk.CTkFrame(control_frame)
            frame.pack(fill="x", padx=20, pady=8)
            ctk.CTkLabel(frame, text=text, width=120, anchor="w").pack(side="left")
            entry = ctk.CTkEntry(frame, textvariable=var, width=100)
            entry.pack(side="right")
            self.entries[text] = entry

        # Gombok
        btn_frame = ctk.CTkFrame(control_frame)
        btn_frame.pack(pady=20)

        self.start_btn = ctk.CTkButton(btn_frame, text="Tanítás indítása", fg_color="green", hover_color="#006400",
                                       command=self.toggle_training, width=200, height=40, font=("Roboto", 14, "bold"))
        self.start_btn.pack(pady=5)

        self.reset_btn = ctk.CTkButton(btn_frame, text="Reset", fg_color="gray", command=self.reset,
                                       width=200, height=40)
        self.reset_btn.pack(pady=5)

        self.export_btn = ctk.CTkButton(btn_frame, text="Mentés Excelbe", fg_color="blue", hover_color="#000080",
                                       command=self.export_to_excel, width=200, height=40)
        self.export_btn.pack(pady=5)

        self.status_label = ctk.CTkLabel(control_frame, text="Állapot: Készen áll", text_color="cyan")
        self.status_label.pack(pady=10)

        self.iter_label = ctk.CTkLabel(control_frame, text="Iteráció: 0", font=("Roboto", 16))
        self.iter_label.pack(pady=5)

        # Táblázat
        self.tree = ttk.Treeview(control_frame, columns=("x1","x2","x3","d","y","net","e"), show="headings", height=20)
        for col in ("x1","x2","x3","d","y","net","e"):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=40, anchor="center")
        self.tree.pack(pady=10, padx=10, fill="both")

        # Táblázat stílus
        style = ttk.Style()
        style.configure("Treeview", background="black", foreground="red", fieldbackground="black", font=('Helvetica', 10, 'bold'))
        style.configure("Treeview.Heading", font=('Helvetica', 11, 'bold'), foreground='red', background='black')

        # === Jobb oldal: 3D plot ===
        plot_frame = ctk.CTkFrame(self.root)
        plot_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        self.fig = Figure(figsize=(10, 8), dpi=100, facecolor='#2b2b2b')
        self.ax = self.fig.add_subplot(111, projection='3d')
        self.ax.set_facecolor('#2b2b2b')
        self.fig.patch.set_facecolor('#2b2b2b')

        self.canvas = FigureCanvasTkAgg(self.fig, plot_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

    def update_plot(self):
        self.ax.cla()

        self.ax.set_xlim(0, 1)
        self.ax.set_ylim(0, 1)
        self.ax.set_zlim(0, 1)
        self.ax.set_xlabel('x1', color='white')
        self.ax.set_ylabel('x2', color='white')
        self.ax.set_zlabel('x3', color='white')
        self.ax.tick_params(colors='white')
        self.ax.grid(True, color='gray', alpha=0.3)

        # Pontok
        for x1, x2, x3, d in self.data:
            self.ax.scatter(x1, x2, x3,
                            c='lime' if d else 'red',
                            marker='^' if d else 'o',
                            s=300 if d else 150,
                            edgecolors='white', linewidth=1.5)

        # Döntési sík
        if abs(self.w3.get()) > 1e-8:
            xx, yy = np.meshgrid(np.linspace(0, 1, 20), np.linspace(0, 1, 20))
            zz = (-self.w1.get() * xx - self.w2.get() * yy - self.b.get()) / self.w3.get()
            zz = np.clip(zz, 0, 1)
            self.ax.plot_surface(xx, yy, zz, color='cyan', alpha=0.5, edgecolor='none')

        self.ax.set_title(
            f'w1={self.w1.get():.3f}  w2={self.w2.get():.3f}  w3={self.w3.get():.3f}  b={self.b.get():.3f}',
            color='cyan', pad=20)

        self.canvas.draw()

    def reset(self):
        self.running = False
        self.converged = False
        self.w1.set(0.7)
        self.w2.set(0.2)
        self.w3.set(0.8)
        self.b.set(0.10)
        self.iteration_records.clear()
        self.update_plot()
        self.iter_label.configure(text="Iteráció: 0")
        self.status_label.configure(text="Állapot: Resetelve")
        for item in self.tree.get_children():
            self.tree.delete(item)

    def toggle_training(self):
        if not self.running:
            self.running = True
            self.start_btn.configure(text="Tanítás leállítása", fg_color="red")
            self.status_label.configure(text="Állapot: Tanítás folyamatban...")
            threading.Thread(target=self.training_loop, daemon=True).start()
        else:
            self.running = False
            self.start_btn.configure(text="Tanítás indítása", fg_color="green")
            self.status_label.configure(text="Állapot: Megállítva")

    def training_loop(self):
        iteration = 0
        self.converged = False

        while self.running and iteration < self.max_iter.get() and not self.converged:
            iteration += 1
            self.iter_label.configure(text=f"Iteráció: {iteration}")

            # Minden sor törlése
            for item in self.tree.get_children():
                self.tree.delete(item)

            misclassified = 0
            current_iter_data = []

            for x1, x2, x3, d in self.data:
                net = self.w1.get()*x1 + self.w2.get()*x2 + self.w3.get()*x3 + self.b.get()
                y = self.step_function(net)
                e = d - y

                tag = "green" if e == 0 else "red"
                self.tree.insert("", "end", values=(x1,x2,x3,d,y,f"{net:.3f}",f"{e:+d}"), tags=(tag,))

                current_iter_data.append((x1,x2,x3,d,y,f"{net:.3f}",f"{e:+d}", tag))

                if e != 0:
                    misclassified += 1
                    self.w1.set(self.w1.get() + self.lr.get()*e*x1)
                    self.w2.set(self.w2.get() + self.lr.get()*e*x2)
                    self.w3.set(self.w3.get() + self.lr.get()*e*x3)
                    self.b.set(self.b.get() + self.lr.get()*e)

            self.tree.tag_configure("green", foreground="#00ff00")
            self.tree.tag_configure("red", foreground="#ff4444")

            self.iteration_records.append(current_iter_data)

            self.update_plot()

            if misclassified == 0:
                self.converged = True
                break

            # Iterációk közti szünet
            time.sleep(0.4)

        self.running = False
        self.start_btn.configure(text="Tanítás indítása", fg_color="green")

        if self.converged:
            self.status_label.configure(text="KONVERGENCIA – Tanítás kész!", text_color="#00ff00")
        else:
            self.status_label.configure(text="Leállt / Max iteráció elérve", text_color="orange")

        self.run_test_phase()

    def run_test_phase(self):
        """Tanítás után automatikus tesztelés"""
        self.tree.insert("", "end", values=("", "", "", "", "", "", ""), tags=("sep",))
        self.tree.tag_configure("sep", foreground="#00ffff")

        for x1, x2, x3, d in self.data:
            net = self.w1.get()*x1 + self.w2.get()*x2 + self.w3.get()*x3 + self.b.get()
            y = self.step_function(net)
            e = d - y

            tag = "test_ok" if e == 0 else "test_bad"
            self.tree.insert("", "end", values=(x1,x2,x3,d,y,f"{net:.3f}",f"{e:+d}"), tags=(tag,))

        self.tree.tag_configure("test_ok", foreground="#00ffaa")
        self.tree.tag_configure("test_bad", foreground="#ff0066")

    def export_to_excel(self):
        """Iterációk mentése Excel fájlba, színezéssel és elválasztó sorokkal"""
        if not self.iteration_records:
            tk.messagebox.showinfo("Info", "Nincs menteni való adat!")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files","*.xlsx")])
        if not file_path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Perceptron Tanítás"

        headers = ["x1","x2","x3","d","y","net","e"]
        ws.append(headers)

        for iter_data in self.iteration_records:
            # Elválasztó sor
            ws.append([""]*len(headers))
            for row in iter_data:
                ws.append(row[:7])
                fill_color = "00ff00" if row[7]=="green" else "ff0000"
                for col_idx in range(1,8):
                    cell = ws.cell(row=ws.max_row, column=col_idx)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                    cell.font = Font(bold=True, color="000000")

        wb.save(file_path)
        tk.messagebox.showinfo("Siker", f"Adatok elmentve: {file_path}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = Perceptron3DApp()
    app.run()
